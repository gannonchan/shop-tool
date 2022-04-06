import resource
import initdata
from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl
import decimal
import datetime
import time
from sheetvo import SheetVo
from dbmgr import DbManager
import conf


def init_db():
    # DbManager.__init__("tablehead.db")
    # sqlquery = DbManager.getsqlquery()
    dbMgr = DbManager("tablehead.db")
    #创建店铺表
    create_shop_sql = """CREATE TABLE IF NOT EXISTS SHOP(
                   ID INTEGER PRIMARY KEY AUTOINCREMENT,
                   SHOP_NAME           TEXT    NOT NULL,
                   ADDRESS             TEXT
                )"""
    #创建店铺表头字典表
    create_shop_column_mapping_sql = """CREATE TABLE IF NOT EXISTS COLUMN_MAPPING(
                       ID INTEGER PRIMARY KEY AUTOINCREMENT,
                       SRC_NAME           TEXT    NOT NULL,
                       TAG_NAME           TEXT    NOT NULL,
                       SHOP_ID            INTEGER NOT NULL,
                       SHOP_NAME          TEXT    NOT NULL
                    )"""
    # sqlquery.exec()
    # conn = sqlite3.connect('tablehead.db')
    conn =  dbMgr.getconn()
    cursor = conn.cursor()
    cursor.execute(create_shop_sql)
    cursor.execute(create_shop_column_mapping_sql)
    conn.commit()
    list_shop_sql = "SELECT * FROM SHOP"
    cursor.execute(list_shop_sql)
    result = cursor.fetchall()
    for row in result:
        (id, shop_name, addr) = row
        list_column_map_sql = "SELECT * FROM COLUMN_MAPPING WHERE SHOP_ID = {}".format(id)
        res = cursor.execute(list_column_map_sql)
        column_dict = {}
        for row in res:
            print(row)
            (id, src_name, tag_name, shop_id, shop_name, date_col_num, title_row_num) = row
            column_dict.setdefault(src_name.strip(),tag_name.strip())
        DbManager.shop_dict_cache.setdefault(shop_name.strip(),column_dict)
    DbManager.close(cursor,conn)
    # cursor.close()
    # conn.close()
    print(DbManager.shop_dict_cache)
    pass

def parse_worksheet(src_workbook_path, date):
    src_workbook = openpyxl.load_workbook(src_workbook_path)
    worksheet_name = 'Sheet{}'.format(date)
    worksheet = src_workbook[worksheet_name]
    if not src_workbook.__contains__(worksheet_name):
        return (False,0,"店铺工作簿中工作表{}不存在".format(worksheet_name))
    row_index = 0
    col_dict = {}
    shop_name = worksheet['B2'].value
    shop_col_map = DbManager.shop_dict_cache.get(shop_name)
    for cIndex in range(0, 26):
        col_dict.setdefault(cIndex + 1, chr(cIndex + 65))
    title_attr_dict_cn_keys = {}
    title_attr_dict_index_keys = {}
    price_attr_dict = {}
    for row in worksheet.rows:
        row_index += 1
        cell_index = 0
        for cell in row:
            cell_index += 1
            value = cell.value
            # print(cell, type(value), value)
            if not value is None:
                if isinstance(value, str) and len(value.strip()) > 0:
                    print('str:',value.strip())
                    index = col_dict.get(cell_index) + str(row_index)
                    #交换表头
                    tag_col_head = shop_col_map.get(value.strip())
                    # title_attr_dict_cn_keys.setdefault(tag_col_head, index)
                    # 用目标表头作为原表头关联坐标
                    title_attr_dict_index_keys.setdefault(index, tag_col_head)
                elif isinstance(value, float):
                    index = col_dict.get(cell_index) + str(row_index)
                    #用金额作为键存储坐标，用于之后偏移到有金额内容的表头位置
                    price_attr_dict.setdefault(value, index)
    final_dict = get_title_index(price_attr_dict, title_attr_dict_index_keys)
    # shop_name = worksheet['B2'].value
    vo = SheetVo(final_dict,date,shop_name)
    src_workbook.close()
    return (True,vo,'success')


# 获取金额头文本下标
def get_title_index(price_attr_dict, title_attr_dict_index_keys):
    final_dict = {}
    for key in price_attr_dict.keys():
        cell_index = price_attr_dict.get(key)
        col_index = cell_index[0]
        row_index = cell_index[1:]
        col_index_ascii = ord(col_index)
        title_index = chr(col_index_ascii - 1) + row_index
        title_text = title_attr_dict_index_keys.get(title_index)
        final_dict.setdefault(title_text, key)
    return final_dict


def padding_sheet(target_workbook_path, src_sheet_vo):
    target_workbook = openpyxl.load_workbook(target_workbook_path)
    shop_name = src_sheet_vo.getshopName()
    target_worksheet = target_workbook[shop_name]
    day_str = src_sheet_vo.getday()
    #设置目标日期列号
    # date_col = src_sheet_vo.getdaycol()
    # day_column = target_worksheet[date_col]
    day_column = target_worksheet['A']
    exec_row_index = 0
    for cell in day_column:
        exec_row_index += 1
        if cell.value == day_str:
            break
    for column in target_worksheet['B':'AA']:
        # print(column)
        row_index = 0
        value = 0
        flag = False
        for cell in column:
            row_index += 1
            #设置目标表头行号
    # title_row = src_sheet_vo.gettitlerow()
    #         if row_index == title_row:
            if row_index == 2:
                #titleRowIndex
                cell_value = cell.value
                print(cell_value.strip())
                dict = src_sheet_vo.getdict()
                if not flag:
                    value = dict.get(cell_value.strip())
                    if value is None:
                        break
                    else:
                        flag = True
            if row_index == exec_row_index:
                cell.value = value
    target_workbook.save(target_workbook_path)
    target_workbook.close()


class Ui_Form(QtWidgets.QWidget):

    def __init__(self):
        super().__init__()
        # self.configBtn.clicked.connect(self.on_configBtn_clicked)
        # self.child_widget = conf.Ui_childWidget()

    def setupUi(self, Form):
        Form.setObjectName("Form")
        # Form.setMinimumSize(330, 600)
        # Form.setMaximumSize(330, 600)
        Form.resize(330, 600)
        self.gridLayout_4 = QtWidgets.QGridLayout(Form)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.gridLayout_3 = QtWidgets.QGridLayout()
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.groupBox = QtWidgets.QGroupBox(Form)
        self.groupBox.setObjectName("groupBox")
        self.gridLayout = QtWidgets.QGridLayout(self.groupBox)
        self.gridLayout.setObjectName("gridLayout")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        self.targetWbEdit = QtWidgets.QLineEdit(self.groupBox)
        self.targetWbEdit.setObjectName("targetWbEdit")
        self.targetWbEdit.setReadOnly(True)
        self.horizontalLayout.addWidget(self.targetWbEdit)
        self.targetBtn = QtWidgets.QPushButton(self.groupBox)
        self.targetBtn.setObjectName("targetBtn")
        self.targetBtn.setStyleSheet(
            "background-color:#00CED1;padding:2px 4px;color:#fff;font-family:'黑体';font-size:12px;font-weight:bold")
        self.horizontalLayout.addWidget(self.targetBtn)
        self.gridLayout.addLayout(self.horizontalLayout, 0, 0, 1, 1)
        self.gridLayout_3.addWidget(self.groupBox, 0, 0, 1, 1)
        self.groupBox_2 = QtWidgets.QGroupBox(Form)
        self.groupBox_2.setObjectName("groupBox_2")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.groupBox_2)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.label_2 = QtWidgets.QLabel(self.groupBox_2)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_2.addWidget(self.label_2)
        self.srcWbEdit = QtWidgets.QLineEdit(self.groupBox_2)
        self.srcWbEdit.setObjectName("srcWbEdit")
        self.srcWbEdit.setReadOnly(True)
        self.horizontalLayout_2.addWidget(self.srcWbEdit)
        self.srcBtn = QtWidgets.QPushButton(self.groupBox_2)
        self.srcBtn.setObjectName("srcBtn")
        self.srcBtn.setStyleSheet(
            "background-color:#00CED1;padding:2px 4px;color:#fff;font-family:'黑体';font-size:12px;font-weight:bold")
        self.horizontalLayout_2.addWidget(self.srcBtn)
        self.gridLayout_2.addLayout(self.horizontalLayout_2, 0, 0, 1, 1)
        self.gridLayout_3.addWidget(self.groupBox_2, 1, 0, 1, 1)

        #日期标签
        self.dateLabel = QtWidgets.QLabel(Form)
        self.dateLabel.setText("当月日期：")
        self.dateLabel.setMaximumWidth(55)
        
        self.spinbox = QtWidgets.QSpinBox(Form)
        self.spinbox.setObjectName("spinbox")
        self.spinbox.setMaximum(31)
        self.spinbox.setMinimum(1)

        self.spinbox_2 = QtWidgets.QSpinBox(Form)
        self.spinbox_2.setObjectName("spinbox_2")
        self.spinbox_2.setMaximum(31)
        self.spinbox_2.setMinimum(1)

        self.spinbox_3 = QtWidgets.QSpinBox(Form)
        self.spinbox_3.setObjectName("spinbox_3")
        self.spinbox_3.setMaximum(31)
        self.spinbox_3.setMinimum(1)

        self.dateSelectLabel = QtWidgets.QLabel(Form)
        self.dateSelectLabel.setText("抽取三日:")

        self.checkbox = QtWidgets.QCheckBox(Form)
        self.checkbox.setObjectName("checkBox")
        self.checkbox.setChecked(False)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.horizontalLayout_3.addWidget(self.dateLabel, QtCore.Qt.AlignLeft)
        self.horizontalLayout_3.addWidget(self.spinbox, QtCore.Qt.AlignLeft)
        self.horizontalLayout_3.addWidget(self.spinbox_2, QtCore.Qt.AlignLeft)
        self.horizontalLayout_3.addWidget(self.spinbox_3, QtCore.Qt.AlignLeft)
        self.horizontalLayout_3.addWidget(self.dateSelectLabel, QtCore.Qt.AlignLeft)
        self.horizontalLayout_3.addWidget(self.checkbox, QtCore.Qt.AlignLeft)
        self.gridLayout_3.addLayout(self.horizontalLayout_3, 2, 0, 1, 1)
        self.plainTextEdit = QtWidgets.QPlainTextEdit(Form)
        self.plainTextEdit.setObjectName("plainTextEdit")
        self.plainTextEdit.setReadOnly(True)
        self.plainTextEdit.appendPlainText("  - - - - - - - - - -操作日志- - - - - - - - - -  ")
        self.plainTextEdit.appendPlainText("  ---------------------------------------------  ")
        self.gridLayout_3.addWidget(self.plainTextEdit, 3, 0, 1, 1)
        self.configBtn = QtWidgets.QPushButton(Form)
        self.configBtn.setObjectName('configBtn')
        self.configBtn.setText('设置')
        self.configBtn.setStyleSheet(
            "QPushButton{background-color: #FF7F50; padding:4px; border:2px solid gray; border-radius:5px; border-color:#fff;"
            " color:#FFF; font-family:'黑体';font-size:12px;font-weight:bold}"
            "QPushButton:hover{background-color:#FFA07A; color: black;}"
            "QPushButton:pressed{background-color:#FF6347;border-style: inset;}")
        self.clearLogBtn = QtWidgets.QPushButton(Form)
        self.clearLogBtn.setObjectName('clearLogBtn')
        self.clearLogBtn.setText('清空操作日志')
        self.clearLogBtn.setStyleSheet("QPushButton{background-color: #00FF00; padding:4px; border:2px solid gray; border-radius:5px; border-color:#fff;"
                                    " color:#FFF; font-family:'黑体';font-size:12px;font-weight:bold}"
                                    "QPushButton:hover{background-color:#66FF66; color: black;}"
                                    "QPushButton:pressed{background-color:#00BB00;border-style: inset;}")
        self.startBtn = QtWidgets.QPushButton(Form)
        self.startBtn.setObjectName("startBtn")
        self.startBtn.setStyleSheet("QPushButton{background-color: #EB1479; padding:4px; border:2px solid gray; border-radius:5px; border-color:#fff;"
                                    " color:#FFF; font-family:'黑体';font-size:12px;font-weight:bold}"
                                    "QPushButton:hover{background-color:#FF69B4; color: black;}"
                                    "QPushButton:pressed{background-color:#C71585;border-style: inset;}")
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName('horizontalLayout_4')
        self.horizontalLayout_4.addWidget(self.configBtn, QtCore.Qt.AlignLeft)
        self.horizontalLayout_4.addWidget(self.clearLogBtn, QtCore.Qt.AlignLeft)
        self.horizontalLayout_4.addWidget(self.startBtn, QtCore.Qt.AlignLeft)
        self.gridLayout_3.addLayout(self.horizontalLayout_4, 4, 0, 1, 1)
        self.copyLable = QtWidgets.QLabel(Form)
        self.copyLable.setText('©版权声明:仅供雷容个人使用,侵权承担法律责任')
        self.gridLayout_3.addWidget(self.copyLable, 5, 0, 1, 1)
        self.gridLayout_4.addLayout(self.gridLayout_3, 0, 0, 1, 1)
        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)


        self.targetBtn.clicked.connect(self.on_targetBtn_clicked)
        self.srcBtn.clicked.connect(self.on_srcBtn_clicked)
        self.clearLogBtn.clicked.connect(self.on_clearLogBtn_clicked)
        self.startBtn.clicked.connect(self.on_startBtn_clicked)
        self.configBtn.clicked.connect(self.on_configBtn_clicked)
        # self.spinbox.valueChanged.connect(self.on_spinbox_valueChanged)
        # self.spinbox_2.valueChanged.connect(self.on_spinbox_2_valueChanged)
        # self.spinbox_3.valueChanged.connect(self.on_spinbox_3_valueChanged)
        # self.checkbox.clicked.connect(self.on_checkbox_clicked)


    def on_configBtn_clicked(self):
        # child_widget = conf.QtWidgets.QWidget()
        # ui = conf.Ui_childWidget()
        # ui.setupUi(child_widget)
        # # widget.setWindowIcon(QtGui.QIcon(":/ico/my.ico"))
        self.child_widget = conf.Ui_childWidget()
        self.child_widget.setupUi(self.child_widget)
        self.child_widget.show()
        pass
    def on_targetBtn_clicked(self):
        (target_wb_path,_) = QtWidgets.QFileDialog.getOpenFileName(None,'选择文件','./','All Files (*.xlsx)')
        self.targetWbEdit.setText(target_wb_path)
        if len(target_wb_path) > 0:
            self.plainTextEdit.appendPlainText("添加了当月汇总工作簿：%s" % target_wb_path)


    def on_srcBtn_clicked(self):
        (src_wb_path,_) = QtWidgets.QFileDialog.getOpenFileName(None,'选择文件','./','All Files (*.xlsx)')
        self.srcWbEdit.setText(src_wb_path)
        if len(src_wb_path) > 0:
            self.plainTextEdit.appendPlainText("添加了当月店铺工作簿：%s" % src_wb_path)


    def on_clearLogBtn_clicked(self):
        self.plainTextEdit.clear()


    def on_startBtn_clicked(self):
        if not self.is_selected():
            QtWidgets.QMessageBox.critical(None,"错误","缺失完整工作簿信息")
        else:
            date = 0
            if self.checkbox.isChecked():
                date = [self.spinbox.value(),self.spinbox_2.value(),self.spinbox_3.value()]
            else:
                date = [self.spinbox.value()]
            self.on_spinbox_valueChanged()
            data = initdata.InitData(self.targetWbEdit.text(), self.srcWbEdit.text(), date)
            self.plainTextEdit.appendPlainText("店铺工作簿{}第{}日的内容将会汇总到工作簿{}".format(data.getsrc_wb_path(),
                          data.getdate(), data.gettarget_wb_path()))
            target_wb_path = data.gettarget_wb_path()
            oldwb = openpyxl.load_workbook(target_wb_path)
            index = target_wb_path.rfind('.xlsx')
            prefix = target_wb_path[0:index]
            suffix = target_wb_path[index:]
            datetimestr = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            back_path = prefix+datetimestr+suffix
            print(back_path)
            oldwb.save(back_path)
            oldwb.close()
            self.plainTextEdit.appendPlainText('备份{}工作簿到{}'.format(target_wb_path,back_path))
            for date in data.getdate():
                (code,vo,msg) = parse_worksheet(data.getsrc_wb_path(), date)
                self.plainTextEdit.appendPlainText("开始导入店铺工作簿{}Sheet{}表"
                                                   .format(data.getsrc_wb_path(),data.getdate()))
                self.plainTextEdit.appendPlainText('请稍后...')
                if code:
                    padding_sheet(data.gettarget_wb_path(), vo)
                else:
                    QtWidgets.QMessageBox.critical(None,"导出错误",msg)
                    return
            self.plainTextEdit.appendPlainText('导出结束!!!')


    # def on_checkbox_clicked(self):
    #     self.on_spinbox_valueChanged()
    #
    #
    def on_spinbox_valueChanged(self):
        if not self.checkbox.isChecked():
            self.plainTextEdit.appendPlainText('导入当月第{}号的店铺数据'.format(self.spinbox.value()))
        else:
            self.plainTextEdit.appendPlainText('导入当月第{}、{}、{}号的店铺数据'
                                               .format(self.spinbox.value(), self.spinbox_2.value(),
                                                       self.spinbox_3.value()))
    #
    # def on_spinbox_2_valueChanged(self):
    #     if self.checkbox.isChecked():
    #         self.plainTextEdit.appendPlainText('导入当月第{}、{}、{}号的店铺数据'
    #                                            .format(self.spinbox.value(),self.spinbox_2.value(),
    #                                                    self.spinbox_3.value()))
    #
    #
    # def on_spinbox_3_valueChanged(self):
    #     if self.checkbox.isChecked():
    #         self.plainTextEdit.appendPlainText('导入当月第{}、{}、{}号的店铺数据'
    #                                            .format(self.spinbox.value(),self.spinbox_2.value(),
    #                                                    self.spinbox_3.value()))


    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "汇总数据小工具"))
        self.groupBox.setTitle(_translate("Form", "请选择数据目标工作簿"))
        self.label.setText(_translate("Form", "汇总工作簿:"))
        self.targetBtn.setText(_translate("Form", "选择工作簿"))
        self.groupBox_2.setTitle(_translate("Form", "请选择数据来源工作簿"))
        self.label_2.setText(_translate("Form", "店铺工作簿:"))
        self.srcBtn.setText(_translate("Form", "选择工作簿"))
        self.startBtn.setText(_translate("Form", "开始抽取"))

    def is_selected(self):
        if len(self.targetWbEdit.text()) > 0 and len(self.srcWbEdit.text()) > 0:
            return True
        else:
            return False


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    widget = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(widget)
    widget.setWindowIcon(QtGui.QIcon(":/ico/my.ico"))
    widget.show()
    init_db()
    sys.exit(app.exec_())